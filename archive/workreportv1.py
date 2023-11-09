import openpyxl
from tkinter import *
from tkinter import ttk
from tkinter import messagebox

# Загрузите данные из файла Excel
workbook = openpyxl.load_workbook('database.xlsx')
sheet = workbook.active

# Создайте приложение Tkinter
root = Tk()

# Создайте поля для ввода данных
fields = ['участок', 'фамилия', 'шифр детали', 'наименование детали', 'количество деталей', 'операция', 'норма времени на единицу']
entries = {}

# Прочитайте все данные из листа в виде списка списков
data = list(sheet.values)

def update_combobox(event):
    # Получить текущий текст
    current_text = event.widget.get()
    # Получить все значения
    values = event.widget["values"]
    # Фильтровать значения
    new_values = [value for value in values if value.startswith(current_text)]
    # Обновить выпадающий список
    event.widget["values"] = new_values

for field in fields:
    # Создайте выпадающий список для каждого поля
    label = Label(root, text=field)
    label.pack()
    combo = ttk.Combobox(root)
    # Используйте индекс для доступа к определенному столбцу
    combo['values'] = list(set([row[fields.index(field)] for row in data]))
    combo.bind("<<ComboboxSelected>>", update_combobox)
    combo.pack()
    entries[field] = combo

# Создайте функцию для сохранения данных
def save_data():
    # Сохраните данные в файлы Excel
    for filename in ['сменное.xlsx', 'отчет.xlsx']:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        row = sheet.max_row + 1
        for field in fields:
            cell = sheet.cell(row=row, column=fields.index(field) + 1)
            cell.value = entries[field].get()
        workbook.save(filename)
    # Показать сообщение об успешном сохранении
    messagebox.showinfo("Сохранение", "Данные успешно сохранены!")

# Создайте кнопку "сохранить"
button = Button(root, text="сохранить", command=save_data)
button.pack()

root.mainloop()
