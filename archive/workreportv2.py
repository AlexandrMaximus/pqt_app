from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QLabel, QLineEdit, QComboBox, QCompleter
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtCore import Qt
import openpyxl

# Загрузите данные из файла Excel
workbook = openpyxl.load_workbook('database.xlsx')
sheet = workbook.active

# Создайте приложение PyQt
app = QApplication([])

# Создайте главное окно
window = QWidget()
layout = QVBoxLayout(window)

# Создайте поля для ввода данных
fields = ['участок', 'фамилия', 'шифр детали', 'наименование детали', 'количество деталей', 'операция', 'норма времени на единицу']
entries = {}

# Прочитайте все данные из листа в виде списка списков
data = list(sheet.values)

for field in fields:
    # Создайте выпадающий список для каждого поля
    label = QLabel(field)
    layout.addWidget(label)
    edit = QLineEdit()
    completer = QCompleter(list(set([str(row[fields.index(field)]) for row in data])))
    completer.setCaseSensitivity(Qt.CaseInsensitive)
    edit.setCompleter(completer)
    layout.addWidget(edit)
    entries[field] = edit

# Создайте функцию для сохранения данных
def save_data():
    # Сохраните данные в файлы Excel
    for filename in ['сменное.xlsx', 'отчет.xlsx']:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        row = sheet.max_row + 1
        for field in fields:
            cell = sheet.cell(row=row, column=fields.index(field) + 1)
            cell.value = entries[field].text()
        workbook.save(filename)
    # Показать сообщение об успешном сохранении
    msgBox = QMessageBox()
    msgBox.setIcon(QMessageBox.Information)
    msgBox.setText("Данные успешно сохранены!")
    msgBox.exec_()

# Создайте кнопку "сохранить"
button = QPushButton('сохранить')
button.clicked.connect(save_data)
layout.addWidget(button)

window.show()
app.exec_()
